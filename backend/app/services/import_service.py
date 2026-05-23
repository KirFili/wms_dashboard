"""
ETL-сервис импорта XLSX из 1С.
Парсинг openpyxl → mapping → validation → staging → upsert.
"""

import hashlib
import uuid
from datetime import date, datetime, timezone
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.import_batch import ImportBatch
from app.models.import_row import ImportRow
from app.models.sku import SKU
from app.models.warehouse import Warehouse
from app.models.chamber import Chamber
from app.models.pallet_location import PalletLocation
from app.models.stock_snapshot import StockSnapshot
from app.models.shipment_fact import ShipmentFact
from app.models.movement_fact import MovementFact

# ── Column name aliases (гибкий маппинг) ──────────

COLUMN_ALIASES: dict[str, list[str]] = {
    "snapshot_date": ["дата", "date", "snapshot_date", "дата_снимка", "дата остатков"],
    "shipment_date": ["дата отгрузки", "shipment_date", "дата_отгрузки"],
    "operation_datetime": ["дата_операции", "operation_datetime", "дата операции", "datetime"],
    "warehouse_code": ["склад", "warehouse", "warehouse_code", "код_склада"],
    "chamber_code": ["камера", "chamber", "chamber_code", "код_камеры"],
    "pallet_code": ["паллетоместо", "pallet", "pallet_code", "ячейка", "location"],
    "sku_code": ["sku", "sku_code", "код_sku", "номенклатура", "код номенклатуры", "артикул"],
    "occupied_locations": ["занято_мест", "occupied_locations", "паллет", "pallets"],
    "occupied_volume": ["объем", "volume", "occupied_volume", "объём"],
    "shipped_qty": ["отгружено", "shipped_qty", "количество", "qty", "кол-во"],
    "shipped_volume": ["объём_отгрузки", "shipped_volume", "объем отгрузки"],
    "document_number": ["номер_документа", "document_number", "документ", "doc_number"],
    "operation_type": ["тип_операции", "operation_type", "операция", "operation"],
    "qty": ["количество", "qty", "кол-во"],
}

# ── Expected columns per import type ────────────────

EXPECTED_COLUMNS = {
    "stock_snapshot": {
        "required": ["snapshot_date", "warehouse_code", "chamber_code", "sku_code", "occupied_locations"],
        "optional": ["pallet_code", "occupied_volume"],
    },
    "shipment": {
        "required": ["shipment_date", "warehouse_code", "sku_code", "shipped_qty"],
        "optional": ["shipped_volume", "document_number"],
    },
    "movement": {
        "required": ["operation_datetime", "warehouse_code", "chamber_code", "operation_type"],
        "optional": ["pallet_code", "sku_code", "qty", "volume"],
    },
}


class ImportService:
    """ETL-сервис импорта XLSX-файлов из 1С."""

    # ═══════════════════════════════════════════════
    # Этап 1: Parse — чтение XLSX
    # ═══════════════════════════════════════════════

    @staticmethod
    async def parse_xlsx(
        db: AsyncSession,
        batch: ImportBatch,
        file_path: str,
    ) -> ImportBatch:
        """Парсинг XLSX-файла: чтение всех листов, mapping колонок, запись строк."""
        wb = load_workbook(file_path, read_only=True, data_only=True)

        all_rows: list[dict[str, Any]] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            # Первая строка — заголовки
            try:
                raw_headers: list[str] = [
                    str(h).strip().lower() if h is not None else ""
                    for h in next(rows_iter)
                ]
            except StopIteration:
                continue

            # Строим маппинг: индекс колонки → внутреннее имя поля
            mapping = ImportService._build_mapping(raw_headers, batch.import_type)

            for data_row in rows_iter:
                if all(v is None for v in data_row):
                    continue  # пропускаем пустые строки
                raw: dict[str, Any] = {
                    str(raw_headers[i]): data_row[i]
                    for i in range(len(data_row))
                    if i < len(raw_headers)
                }
                mapped: dict[str, Any] = {}
                for idx, field_name in mapping.items():
                    if idx < len(data_row):
                        value = data_row[idx]
                        if value is not None:
                            mapped[field_name] = str(value).strip()

                all_rows.append({"raw": raw, "mapped": mapped})

        wb.close()

        # Запись строк в staging
        batch.total_rows = len(all_rows)
        for row_num, row_data in enumerate(all_rows):
            import_row = ImportRow(
                batch_id=batch.id,
                row_number=row_num + 1,
                raw_data=row_data["raw"],
                mapped_data=row_data["mapped"],
                validation_status="pending",
                created_at=datetime.now(timezone.utc),
            )
            db.add(import_row)

        batch.status = "parsed"
        await db.flush()
        return batch

    @staticmethod
    def _build_mapping(
        headers: list[str], import_type: str
    ) -> dict[int, str]:
        """Построить маппинг: индекс колонки → внутреннее имя поля."""
        mapping: dict[int, str] = {}
        for idx, header in enumerate(headers):
            h_lower = header.strip().lower()
            for field_name, aliases in COLUMN_ALIASES.items():
                if h_lower in [a.lower() for a in aliases]:
                    mapping[idx] = field_name
                    break
        return mapping

    # ═══════════════════════════════════════════════
    # Этап 2: Validate — проверка данных
    # ═══════════════════════════════════════════════

    @staticmethod
    async def validate_batch(
        db: AsyncSession,
        batch: ImportBatch,
    ) -> ImportBatch:
        """Валидация всех строк пакета импорта."""
        result = await db.execute(
            select(ImportRow)
            .where(ImportRow.batch_id == batch.id)
            .order_by(ImportRow.row_number)
        )
        rows = list(result.scalars().all())

        required = EXPECTED_COLUMNS.get(batch.import_type, {}).get("required", [])

        total_success = 0
        total_warning = 0
        total_error = 0

        # Предзагрузка справочников для проверки существования
        sku_codes_cache = await ImportService._cache_sku_codes(db)
        warehouse_codes_cache = await ImportService._cache_warehouse_codes(db)
        chamber_codes_cache = await ImportService._cache_chamber_codes(db)

        for row in rows:
            errors: list[str] = []
            warnings: list[str] = []
            mapped = row.mapped_data or {}

            # Проверка обязательных полей
            for req_field in required:
                if req_field not in mapped or not mapped.get(req_field):
                    errors.append(f"Missing required field: {req_field}")

            # Проверка дат
            if "snapshot_date" in mapped:
                try:
                    ImportService._parse_date(mapped["snapshot_date"])
                except ValueError as e:
                    errors.append(f"Invalid snapshot_date: {e}")

            if "shipment_date" in mapped:
                try:
                    ImportService._parse_date(mapped["shipment_date"])
                except ValueError as e:
                    errors.append(f"Invalid shipment_date: {e}")

            if "operation_datetime" in mapped:
                try:
                    ImportService._parse_datetime(mapped["operation_datetime"])
                except ValueError as e:
                    errors.append(f"Invalid operation_datetime: {e}")

            # Проверка числовых полей
            for num_field in [
                "occupied_locations", "occupied_volume", "shipped_qty",
                "shipped_volume", "qty",
            ]:
                if num_field in mapped:
                    try:
                        val = float(mapped[num_field])
                        if val < 0:
                            errors.append(f"Negative value for {num_field}: {val}")
                    except (ValueError, TypeError):
                        errors.append(f"Non-numeric value for {num_field}: {mapped[num_field]}")

            # Проверка существования SKU
            if "sku_code" in mapped and mapped["sku_code"]:
                if mapped["sku_code"] not in sku_codes_cache:
                    warnings.append(f"SKU not found: {mapped['sku_code']} (will be skipped)")

            # Проверка существования склада
            if "warehouse_code" in mapped and mapped["warehouse_code"]:
                if mapped["warehouse_code"] not in warehouse_codes_cache:
                    errors.append(f"Warehouse not found: {mapped['warehouse_code']}")

            # Проверка существования камеры
            if "chamber_code" in mapped and mapped["chamber_code"]:
                if mapped["chamber_code"] not in chamber_codes_cache:
                    errors.append(f"Chamber not found: {mapped['chamber_code']}")

            # Генерация natural key
            row.natural_key = ImportService._build_natural_key(mapped, batch.import_type)

            # Итоговый статус
            if errors:
                row.validation_status = "error"
                row.error_messages = {"errors": errors, "warnings": warnings}
                total_error += 1
            elif warnings:
                row.validation_status = "warning"
                row.error_messages = {"warnings": warnings}
                total_warning += 1
            else:
                row.validation_status = "valid"
                total_success += 1

        batch.success_rows = total_success
        batch.warning_rows = total_warning
        batch.error_rows = total_error
        batch.validated_at = datetime.now(timezone.utc)

        if total_error > 0 and total_success == 0:
            batch.status = "error"
        elif total_warning > 0:
            batch.status = "warning"
        else:
            batch.status = "validated"

        await db.flush()
        return batch

    @staticmethod
    async def _cache_sku_codes(db: AsyncSession) -> set[str]:
        result = await db.execute(select(SKU.sku_code).where(SKU.is_active == True))
        return {row[0] for row in result.fetchall()}

    @staticmethod
    async def _cache_warehouse_codes(db: AsyncSession) -> set[str]:
        result = await db.execute(
            select(Warehouse.code).where(Warehouse.is_active == True)
        )
        return {row[0] for row in result.fetchall()}

    @staticmethod
    async def _cache_chamber_codes(db: AsyncSession) -> set[str]:
        result = await db.execute(
            select(Chamber.code).where(Chamber.is_active == True)
        )
        return {row[0] for row in result.fetchall()}

    # ═══════════════════════════════════════════════
    # Этап 3: Commit — upsert в production-таблицы
    # ═══════════════════════════════════════════════

    @staticmethod
    async def commit_batch(
        db: AsyncSession,
        batch: ImportBatch,
    ) -> ImportBatch:
        """Записать валидные строки в production-таблицы (upsert)."""

        if batch.status not in ("validated", "warning"):
            raise ValueError(f"Cannot commit batch in status '{batch.status}'")

        result = await db.execute(
            select(ImportRow)
            .where(
                ImportRow.batch_id == batch.id,
                ImportRow.validation_status.in_(["valid", "warning"]),
            )
            .order_by(ImportRow.row_number)
        )
        rows = list(result.scalars().all())

        # Предзагрузка ID справочников
        sku_map = await ImportService._cache_sku_map(db)
        warehouse_map = await ImportService._cache_warehouse_map(db)
        chamber_map = await ImportService._cache_chamber_map(db)
        pallet_map = await ImportService._cache_pallet_map(db)

        committed = 0
        for row in rows:
            mapped = row.mapped_data or {}
            nk = row.natural_key or ""

            if batch.import_type == "stock_snapshot":
                commit_ok = await ImportService._upsert_stock_snapshot(
                    db, batch, mapped, nk, sku_map, warehouse_map, chamber_map, pallet_map
                )
            elif batch.import_type == "shipment":
                commit_ok = await ImportService._upsert_shipment_fact(
                    db, batch, mapped, nk, sku_map, warehouse_map
                )
            elif batch.import_type == "movement":
                commit_ok = await ImportService._upsert_movement_fact(
                    db, batch, mapped, nk, sku_map, warehouse_map, chamber_map, pallet_map
                )
            else:
                commit_ok = False

            if commit_ok:
                committed += 1

        batch.status = "imported"
        batch.imported_at = datetime.now(timezone.utc)
        batch.success_rows = committed
        await db.flush()
        return batch

    @staticmethod
    async def _upsert_stock_snapshot(
        db: AsyncSession,
        batch: ImportBatch,
        mapped: dict,
        nk: str,
        sku_map: dict[str, uuid.UUID],
        warehouse_map: dict[str, uuid.UUID],
        chamber_map: dict[str, uuid.UUID],
        pallet_map: dict[str, uuid.UUID],
    ) -> bool:
        sku_id = sku_map.get(mapped.get("sku_code", ""))
        wh_id = warehouse_map.get(mapped.get("warehouse_code", ""))
        ch_id = chamber_map.get(mapped.get("chamber_code", ""))
        pl_id = pallet_map.get(mapped.get("pallet_code", ""))

        if not (sku_id and wh_id and ch_id):
            return False

        snap_date = ImportService._parse_date(mapped.get("snapshot_date", ""))
        occ_locs = int(float(mapped.get("occupied_locations", 0)))
        occ_vol = None
        if "occupied_volume" in mapped:
            occ_vol = float(mapped.get("occupied_volume", 0))

        await db.execute(
            text("""
                INSERT INTO wms.stock_snapshots (
                    snapshot_date, warehouse_id, chamber_id, pallet_location_id,
                    sku_id, occupied_locations, occupied_volume,
                    source_batch_id, natural_key
                ) VALUES (
                    :snapshot_date, :warehouse_id, :chamber_id, :pallet_location_id,
                    :sku_id, :occupied_locations, :occupied_volume,
                    :source_batch_id, :natural_key
                )
                ON CONFLICT (natural_key) DO UPDATE SET
                    occupied_locations = EXCLUDED.occupied_locations,
                    occupied_volume = EXCLUDED.occupied_volume,
                    source_batch_id = EXCLUDED.source_batch_id
            """),
            {
                "snapshot_date": snap_date,
                "warehouse_id": wh_id,
                "chamber_id": ch_id,
                "pallet_location_id": pl_id if pl_id else None,
                "sku_id": sku_id,
                "occupied_locations": occ_locs,
                "occupied_volume": occ_vol,
                "source_batch_id": batch.id,
                "natural_key": nk,
            },
        )
        return True

    @staticmethod
    async def _upsert_shipment_fact(
        db: AsyncSession,
        batch: ImportBatch,
        mapped: dict,
        nk: str,
        sku_map: dict[str, uuid.UUID],
        warehouse_map: dict[str, uuid.UUID],
    ) -> bool:
        sku_id = sku_map.get(mapped.get("sku_code", ""))
        wh_id = warehouse_map.get(mapped.get("warehouse_code", ""))

        if not (sku_id and wh_id):
            return False

        ship_date = ImportService._parse_date(mapped.get("shipment_date", ""))
        sq = float(mapped.get("shipped_qty", 0))
        sv = None
        if "shipped_volume" in mapped:
            sv = float(mapped.get("shipped_volume", 0))
        doc = mapped.get("document_number")

        await db.execute(
            text("""
                INSERT INTO wms.shipment_facts (
                    shipment_date, warehouse_id, sku_id,
                    shipped_qty, shipped_volume, document_number,
                    source_batch_id, natural_key
                ) VALUES (
                    :shipment_date, :warehouse_id, :sku_id,
                    :shipped_qty, :shipped_volume, :document_number,
                    :source_batch_id, :natural_key
                )
                ON CONFLICT (natural_key) DO UPDATE SET
                    shipped_qty = EXCLUDED.shipped_qty,
                    shipped_volume = EXCLUDED.shipped_volume,
                    document_number = EXCLUDED.document_number,
                    source_batch_id = EXCLUDED.source_batch_id
            """),
            {
                "shipment_date": ship_date,
                "warehouse_id": wh_id,
                "sku_id": sku_id,
                "shipped_qty": sq,
                "shipped_volume": sv,
                "document_number": doc,
                "source_batch_id": batch.id,
                "natural_key": nk,
            },
        )
        return True

    @staticmethod
    async def _upsert_movement_fact(
        db: AsyncSession,
        batch: ImportBatch,
        mapped: dict,
        nk: str,
        sku_map: dict[str, uuid.UUID],
        warehouse_map: dict[str, uuid.UUID],
        chamber_map: dict[str, uuid.UUID],
        pallet_map: dict[str, uuid.UUID],
    ) -> bool:
        wh_id = warehouse_map.get(mapped.get("warehouse_code", ""))
        ch_id = chamber_map.get(mapped.get("chamber_code", ""))

        if not (wh_id and ch_id):
            return False

        op_dt = ImportService._parse_datetime(
            mapped.get("operation_datetime", "")
        )
        sku_id = sku_map.get(mapped.get("sku_code", ""))
        pl_id = pallet_map.get(mapped.get("pallet_code", ""))
        op_type = mapped.get("operation_type", "other")
        qty = None
        if "qty" in mapped:
            qty = float(mapped["qty"])
        vol = None
        if "volume" in mapped:
            vol = float(mapped["volume"])

        await db.execute(
            text("""
                INSERT INTO wms.movement_facts (
                    operation_datetime, warehouse_id, chamber_id,
                    pallet_location_id, sku_id, operation_type,
                    qty, volume, source_batch_id, natural_key
                ) VALUES (
                    :operation_datetime, :warehouse_id, :chamber_id,
                    :pallet_location_id, :sku_id, :operation_type,
                    :qty, :volume, :source_batch_id, :natural_key
                )
                ON CONFLICT (natural_key) DO UPDATE SET
                    qty = EXCLUDED.qty,
                    volume = EXCLUDED.volume,
                    source_batch_id = EXCLUDED.source_batch_id
            """),
            {
                "operation_datetime": op_dt,
                "warehouse_id": wh_id,
                "chamber_id": ch_id,
                "pallet_location_id": pl_id if pl_id else None,
                "sku_id": sku_id if sku_id else None,
                "operation_type": op_type,
                "qty": qty,
                "volume": vol,
                "source_batch_id": batch.id,
                "natural_key": nk,
            },
        )
        return True

    # ═══════════════════════════════════════════════
    # Этап 4: Rollback — откат импорта
    # ═══════════════════════════════════════════════

    @staticmethod
    async def rollback_batch(
        db: AsyncSession,
        batch: ImportBatch,
    ) -> ImportBatch:
        """Откат данных импорта из production-таблиц."""
        if batch.status != "imported":
            raise ValueError(f"Cannot rollback batch in status '{batch.status}'")

        table_map = {
            "stock_snapshot": "wms.stock_snapshots",
            "shipment": "wms.shipment_facts",
            "movement": "wms.movement_facts",
        }
        table = table_map.get(batch.import_type)
        if table:
            await db.execute(
                text(f"DELETE FROM {table} WHERE source_batch_id = :batch_id"),
                {"batch_id": batch.id},
            )

        batch.status = "rolled_back"
        await db.flush()
        return batch

    # ═══════════════════════════════════════════════
    # Helpers
    # ═══════════════════════════════════════════════

    @staticmethod
    def _parse_date(value: str) -> date:
        """Парсинг даты из строки (поддержка форматов DD.MM.YYYY, YYYY-MM-DD, DD/MM/YYYY)."""
        value = value.strip()
        for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y.%m.%d"]:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        # Пробуем ISO
        return date.fromisoformat(value)

    @staticmethod
    def _parse_datetime(value: str) -> datetime:
        """Парсинг даты/времени из строки."""
        value = value.strip()
        for fmt in [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S",
            "%d.%m.%Y %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
            "%Y-%m-%d",
        ]:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
        return datetime.fromisoformat(value)

    @staticmethod
    def _build_natural_key(mapped: dict, import_type: str) -> str:
        """Построить natural key согласно §9.5 ТЗ."""
        parts: list[str] = []
        if import_type == "stock_snapshot":
            parts = [
                mapped.get("snapshot_date", ""),
                mapped.get("warehouse_code", ""),
                mapped.get("chamber_code", ""),
                mapped.get("pallet_code", "n/a"),
                mapped.get("sku_code", ""),
            ]
        elif import_type == "shipment":
            parts = [
                mapped.get("shipment_date", ""),
                mapped.get("warehouse_code", ""),
                mapped.get("sku_code", ""),
                mapped.get("document_number", "n/a"),
            ]
        elif import_type == "movement":
            parts = [
                mapped.get("operation_datetime", ""),
                mapped.get("warehouse_code", ""),
                mapped.get("chamber_code", ""),
                mapped.get("pallet_code", "n/a"),
                mapped.get("sku_code", "n/a"),
                mapped.get("operation_type", ""),
            ]
        return "::".join(str(p).strip() for p in parts)

    @staticmethod
    async def _cache_sku_map(db: AsyncSession) -> dict[str, uuid.UUID]:
        result = await db.execute(select(SKU.sku_code, SKU.id))
        return {row[0]: row[1] for row in result.fetchall()}

    @staticmethod
    async def _cache_warehouse_map(db: AsyncSession) -> dict[str, uuid.UUID]:
        result = await db.execute(
            select(Warehouse.code, Warehouse.id).where(Warehouse.is_active == True)
        )
        return {row[0]: row[1] for row in result.fetchall()}

    @staticmethod
    async def _cache_chamber_map(db: AsyncSession) -> dict[str, uuid.UUID]:
        result = await db.execute(
            select(Chamber.code, Chamber.id).where(Chamber.is_active == True)
        )
        return {row[0]: row[1] for row in result.fetchall()}

    @staticmethod
    async def _cache_pallet_map(db: AsyncSession) -> dict[str, uuid.UUID]:
        result = await db.execute(select(PalletLocation.code, PalletLocation.id))
        return {row[0]: row[1] for row in result.fetchall()}

    @staticmethod
    def compute_checksum(file_path: str) -> str:
        """Вычислить SHA-256 контрольную сумму файла."""
        sha = hashlib.sha256()
        with open(file_path, "rb") as f:
            while chunk := f.read(8192):
                sha.update(chunk)
        return sha.hexdigest()
